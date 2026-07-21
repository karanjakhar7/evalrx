from __future__ import annotations

import hashlib
import json
import re
import statistics
import tomllib
import wave
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import EvalConfig
from .jsonl import write_jsonl
from .models import AudioMetadata, CallRecord, DataQuality, SourceType, Turn


TRANSCRIPT_WORKBOOK = Path("Transcripts + Calls/Call_Transcripts_redacted.xlsx")
AUDIO_DIR = Path("Transcripts + Calls/Sample Calls")
DEEPGRAM_DIR = Path("transcripts")
TURN_PATTERN = re.compile(r"^(Agent|User):\s*(.*)$")
REDACTION_PATTERN = re.compile(r"\[[A-Z][A-Z0-9 _-]*\]")


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_labeled_transcript(text: str) -> list[Turn]:
    turns: list[Turn] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = TURN_PATTERN.match(line)
        if match:
            label, content = match.groups()
            if not content.strip():
                continue
            role = "agent" if label == "Agent" else "counterparty"
            turns.append(
                Turn(
                    turn_id=len(turns) + 1,
                    speaker=label,
                    role=role,
                    text=content.strip(),
                )
            )
        elif turns:
            previous = turns[-1]
            turns[-1] = previous.model_copy(update={"text": f"{previous.text} {line}"})
        else:
            turns.append(
                Turn(turn_id=1, speaker="Unknown", role="unknown", text=line)
            )
    return turns


def transcript_quality(text: str, *, audio_word_count: int | None = None) -> str:
    if audio_word_count is not None and audio_word_count < 20:
        return "poor"
    if len(text) < 250:
        return "poor"
    if len(text) < 500:
        return "usable"
    return "good"


def apply_role_overrides(
    turns: list[Turn],
    role_mapping: dict[str, str],
    overrides: dict[str, str],
) -> tuple[list[Turn], dict[str, str], float | None]:
    allowed = {"agent", "counterparty", "unknown"}
    unknown_speakers = set(overrides) - set(role_mapping)
    if unknown_speakers:
        raise ValueError(f"Role override contains unknown speakers: {sorted(unknown_speakers)}")
    invalid_roles = {role for role in overrides.values() if role not in allowed}
    if invalid_roles:
        raise ValueError(f"Role override contains invalid roles: {sorted(invalid_roles)}")
    updated_mapping = {**role_mapping, **overrides}
    updated_turns = [
        turn.model_copy(update={"role": updated_mapping.get(turn.speaker, "unknown")})
        for turn in turns
    ]
    return updated_turns, updated_mapping, 1.0 if overrides else None


def normalize_workbook(config: EvalConfig) -> list[CallRecord]:
    path = config.project_root / TRANSCRIPT_WORKBOOK
    if not path.exists():
        raise FileNotFoundError(f"Transcript workbook not found: {path}")
    workbook_sha = sha256_file(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Transcripts" not in workbook.sheetnames:
        raise ValueError("Expected a 'Transcripts' worksheet")
    sheet = workbook["Transcripts"]
    header = sheet["A1"].value
    if header != "transcript_redacted":
        raise ValueError(f"Unexpected workbook header: {header!r}")

    records: list[CallRecord] = []
    for index, row_number in enumerate(range(2, 52), start=1):
        value = sheet.cell(row=row_number, column=1).value
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"Missing transcript at source row {row_number}")
        text = value.strip()
        turns = parse_labeled_transcript(text)
        if not turns:
            raise ValueError(f"No turns parsed at source row {row_number}")
        notes: list[str] = [
            "Source may contain intentional anonymization cuts; never score them as agent failures."
        ]
        if len(text) < 500:
            notes.append("Short source transcript; review for intentional cut or limited interaction.")
        quality = transcript_quality(text)
        record_hash = sha256_bytes(f"{workbook_sha}:{row_number}:{text}".encode())
        records.append(
            CallRecord(
                call_id=f"transcript_{index:03d}",
                source_call_id=f"workbook_row_{row_number}",
                source_type=SourceType.TRANSCRIPT,
                source_path=str(TRANSCRIPT_WORKBOOK),
                source_row=row_number,
                source_sha256=record_hash,
                transcript=text,
                turns=turns,
                data_quality=DataQuality(
                    transcript_quality=quality,
                    redaction_present=True,
                    truncated=True,
                    missing_operational_context=True,
                    notes=notes,
                ),
                role_mapping={"Agent": "agent", "User": "counterparty"},
                role_mapping_confidence=1.0,
            )
        )
    workbook.close()
    return records


def _deepgram_parts(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    results = payload.get("results") or {}
    utterances = results.get("utterances") or []
    channels = results.get("channels") or []
    alternatives = channels[0].get("alternatives", []) if channels else []
    words = alternatives[0].get("words", []) if alternatives else []
    return utterances, words


def normalize_audio(config: EvalConfig) -> list[CallRecord]:
    audio_dir = config.project_root / AUDIO_DIR
    transcript_dir = config.project_root / DEEPGRAM_DIR
    wav_files = sorted(audio_dir.glob("*.wav"))
    if len(wav_files) != 10:
        raise ValueError(f"Expected 10 WAV files, found {len(wav_files)}")
    override_path = config.path("role_overrides")
    if override_path.exists():
        with override_path.open("rb") as handle:
            role_overrides = tomllib.load(handle)
    else:
        role_overrides = {}

    records: list[CallRecord] = []
    for index, wav_path in enumerate(wav_files, start=1):
        source_id = wav_path.stem.removesuffix("_redacted")
        json_path = transcript_dir / f"{source_id}.json"
        if not json_path.exists():
            raise FileNotFoundError(f"Deepgram response not found: {json_path}")
        payload = json.loads(json_path.read_text())
        utterances, words = _deepgram_parts(payload)

        turns: list[Turn] = []
        for utterance in utterances:
            text = str(utterance.get("transcript") or "").strip()
            if not text:
                continue
            speaker = utterance.get("speaker")
            speaker_label = f"Speaker {speaker}" if speaker is not None else "Speaker ?"
            turns.append(
                Turn(
                    turn_id=len(turns) + 1,
                    speaker=speaker_label,
                    role="unknown",
                    text=text,
                    start_seconds=utterance.get("start"),
                    end_seconds=utterance.get("end"),
                )
            )
        if not turns:
            text = " ".join(str(word.get("punctuated_word") or word.get("word") or "") for word in words)
            text = text.strip()
            if text:
                turns = [Turn(turn_id=1, speaker="Speaker ?", role="unknown", text=text)]

        transcript = "\n".join(f"{turn.speaker}: {turn.text}" for turn in turns)
        speakers = sorted(
            {
                word.get("speaker")
                for word in words
                if word.get("speaker") is not None
            }
        )
        confidences = [
            float(word["confidence"])
            for word in words
            if isinstance(word.get("confidence"), (int, float))
        ]
        with wave.open(str(wav_path), "rb") as audio_file:
            duration = audio_file.getnframes() / audio_file.getframerate()
            channels = audio_file.getnchannels()
            sample_rate = audio_file.getframerate()
            sample_width = audio_file.getsampwidth()

        collapsed = len(speakers) <= 1 and duration > 60
        short_asr = len(words) < 20
        notes: list[str] = [
            "Recording is redacted and may contain intentional anonymization cuts; never score them as agent failures."
        ]
        if collapsed:
            notes.append("Diarization collapsed to one speaker; role assignment requires review.")
        if len(speakers) > 2:
            notes.append("More than two speakers detected; may include transfer, IVR, or diarization noise.")
        if short_asr:
            notes.append("ASR produced fewer than 20 words for a recording over one minute.")
        combined_hash = sha256_bytes(
            f"{sha256_file(wav_path)}:{sha256_file(json_path)}".encode()
        )
        call_id = f"audio_{index:03d}"
        turns, role_mapping, role_mapping_confidence = apply_role_overrides(
            turns,
            {f"Speaker {speaker}": "unknown" for speaker in speakers},
            role_overrides.get(call_id, {}),
        )
        if role_overrides.get(call_id):
            notes.append("Human-reviewed speaker-role override applied.")
        records.append(
            CallRecord(
                call_id=call_id,
                source_call_id=source_id,
                source_type=SourceType.AUDIO,
                source_path=str(wav_path.relative_to(config.project_root)),
                source_sha256=combined_hash,
                transcript=transcript,
                turns=turns,
                data_quality=DataQuality(
                    transcript_quality=transcript_quality(
                        transcript, audio_word_count=len(words)
                    ),
                    redaction_present=True,
                    truncated=True,
                    likely_asr_errors=short_asr,
                    speaker_label_noise=collapsed or len(speakers) > 2,
                    diarization_collapsed=collapsed,
                    more_than_two_speakers=len(speakers) > 2,
                    short_asr_transcript=short_asr,
                    missing_operational_context=True,
                    notes=notes,
                ),
                audio=AudioMetadata(
                    duration_seconds=round(duration, 3),
                    channels=channels,
                    sample_rate_hz=sample_rate,
                    sample_width_bytes=sample_width,
                    word_count=len(words),
                    utterance_count=len(utterances),
                    speaker_count=len(speakers),
                    average_word_confidence=(
                        round(statistics.mean(confidences), 6) if confidences else None
                    ),
                    wav_path=str(wav_path.relative_to(config.project_root)),
                    deepgram_json_path=str(json_path.relative_to(config.project_root)),
                ),
                role_mapping=role_mapping,
                role_mapping_confidence=role_mapping_confidence,
            )
        )
    return records


def prepare_dataset(config: EvalConfig) -> list[CallRecord]:
    records = normalize_workbook(config) + normalize_audio(config)
    call_ids = [record.call_id for record in records]
    if len(call_ids) != len(set(call_ids)):
        raise ValueError("Duplicate normalized call IDs")
    if len(records) != 60:
        raise ValueError(f"Expected 60 normalized calls, found {len(records)}")
    output_path = config.path("normalized")
    write_jsonl(output_path, records)
    return records
