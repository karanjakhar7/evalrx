from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .config import EvalConfig
from .jsonl import read_models, write_jsonl
from .models import (
    CallRecord,
    ClassificationResult,
    HumanLabel,
    SourceType,
    StageRecord,
)


AP_FIELDS = [
    "identity_and_authorization",
    "intent_and_routing",
    "information_capture_and_groundedness",
    "workflow_execution",
    "resolution_and_automation",
    "recovery_and_escalation",
]
PX_FIELDS = [
    "listening_and_comprehension",
    "caller_effort_and_repetition",
    "clarity_and_coherence",
    "trust_and_transparency",
    "closure_and_next_steps",
    "empathy_and_tone",
]
SAFETY_FIELDS = [
    "privacy_and_identity",
    "clinical_safety",
    "fabrication_or_false_confirmation",
    "financial_safety",
    "ai_transparency",
    "call_control",
]


def _feature_set(call: CallRecord, classification: ClassificationResult) -> set[str]:
    length = len(call.transcript)
    length_bucket = "short" if length < 500 else "long" if length > 2500 else "medium"
    risk_terms = {
        term
        for term in ["payment", "medication", "prescription", "urgent", "transfer", "voicemail"]
        if term in call.transcript.casefold()
    }
    quality_flags = {
        key
        for key, value in call.data_quality.model_dump().items()
        if isinstance(value, bool) and value
    }
    return {
        f"workflow:{classification.workflow}",
        f"interaction:{classification.interaction_type}",
        f"disposition:{classification.final_disposition}",
        f"length:{length_bucket}",
        *(f"risk:{term}" for term in risk_terms),
        *(f"quality:{flag}" for flag in quality_flags),
    }


def _greedy_diverse(
    candidates: list[tuple[CallRecord, ClassificationResult]], count: int
) -> list[str]:
    selected: list[str] = []
    covered: set[str] = set()
    remaining = list(candidates)
    while remaining and len(selected) < count:
        scored = []
        for call, classification in remaining:
            features = _feature_set(call, classification)
            novelty = len(features - covered)
            risk_bonus = sum(1 for feature in features if feature.startswith("risk:"))
            quality_bonus = sum(1 for feature in features if feature.startswith("quality:"))
            scored.append((novelty, risk_bonus, quality_bonus, -len(call.transcript), call.call_id))
        scored.sort(reverse=True)
        chosen_id = scored[0][-1]
        call, classification = next(item for item in remaining if item[0].call_id == chosen_id)
        selected.append(chosen_id)
        covered.update(_feature_set(call, classification))
        remaining = [item for item in remaining if item[0].call_id != chosen_id]
    return selected


def select_calibration(config: EvalConfig, run_id: str, calls: list[CallRecord]) -> dict[str, Any]:
    classification_path = config.path("runs") / run_id / "classification.jsonl"
    records = read_models(classification_path, StageRecord)
    classifications = {
        row.metadata.call_id: ClassificationResult.model_validate(row.result)
        for row in records
        if row.result
    }
    transcript_candidates = [
        (call, classifications[call.call_id])
        for call in calls
        if call.source_type == SourceType.TRANSCRIPT and call.call_id in classifications
    ]
    audio_candidates = [
        (call, classifications[call.call_id])
        for call in calls
        if call.source_type == SourceType.AUDIO and call.call_id in classifications
    ]
    transcript_count = int(config.raw["audit"]["calibration_transcripts"])
    development_count = int(config.raw["audit"]["development_transcripts"])
    audio_development_count = int(config.raw["audit"]["audio_development_calls"])
    selected_transcripts = _greedy_diverse(transcript_candidates, transcript_count)
    selected_audio = _greedy_diverse(audio_candidates, len(audio_candidates))
    manifest = {
        "run_id": run_id,
        "selection_method": "deterministic greedy maximum feature coverage",
        "transcript_development": selected_transcripts[:development_count],
        "transcript_holdout": selected_transcripts[development_count:],
        "audio_development": selected_audio[:audio_development_count],
        "audio_validation": selected_audio[audio_development_count:],
    }
    path = config.path("runs") / run_id / "calibration_manifest.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(manifest, indent=2, sort_keys=True))
    return manifest


def select_calibration_offline(
    config: EvalConfig, run_id: str, calls: list[CallRecord]
) -> dict[str, Any]:
    """Create a provisional diverse set when live classification is unavailable."""
    pseudo: dict[str, ClassificationResult] = {
        call.call_id: ClassificationResult(
            call_id=call.call_id,
            direction="unknown",
            workflow="unknown",
            interaction_type="unclear",
            counterparty_type="unknown",
            primary_intent="unknown",
            final_disposition="indeterminate",
            transcript_quality=call.data_quality.transcript_quality,
            confidence=0.0,
            requires_human_review=True,
            reasoning_summary="Offline bootstrap only; replace with live classification.",
        )
        for call in calls
    }
    transcript_candidates = [
        (call, pseudo[call.call_id])
        for call in calls
        if call.source_type == SourceType.TRANSCRIPT
    ]
    audio_candidates = [
        (call, pseudo[call.call_id]) for call in calls if call.source_type == SourceType.AUDIO
    ]
    transcript_count = int(config.raw["audit"]["calibration_transcripts"])
    development_count = int(config.raw["audit"]["development_transcripts"])
    audio_development_count = int(config.raw["audit"]["audio_development_calls"])
    selected_transcripts = _greedy_diverse(transcript_candidates, transcript_count)
    selected_audio = _greedy_diverse(audio_candidates, len(audio_candidates))
    manifest = {
        "run_id": run_id,
        "selection_method": (
            "provisional offline feature coverage; rerun full classification before calibration"
        ),
        "transcript_development": selected_transcripts[:development_count],
        "transcript_holdout": selected_transcripts[development_count:],
        "audio_development": selected_audio[:audio_development_count],
        "audio_validation": selected_audio[audio_development_count:],
    }
    path = config.path("runs") / run_id / "calibration_manifest.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(manifest, indent=2, sort_keys=True))
    return manifest


def calibration_call_ids(manifest: dict[str, Any]) -> list[str]:
    return [
        *manifest["transcript_development"],
        *manifest["transcript_holdout"],
        *manifest["audio_development"],
        *manifest["audio_validation"],
    ]


def _split_by_id(manifest: dict[str, Any]) -> dict[str, str]:
    result: dict[str, str] = {}
    for key, split in [
        ("transcript_development", "development"),
        ("transcript_holdout", "holdout"),
        ("audio_development", "audio_development"),
        ("audio_validation", "audio_validation"),
    ]:
        for call_id in manifest[key]:
            result[call_id] = split
    return result


def _stage_results(config: EvalConfig, run_id: str, stage: str) -> dict[str, StageRecord]:
    path = config.path("runs") / run_id / f"{stage}.jsonl"
    return {record.metadata.call_id: record for record in read_models(path, StageRecord)}


def export_calibration(config: EvalConfig, run_id: str, calls: list[CallRecord]) -> tuple[Path, Path]:
    manifest_path = config.path("runs") / run_id / "calibration_manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError("Run classification and calibration selection first")
    manifest = json.loads(manifest_path.read_text())
    split_by_id = _split_by_id(manifest)
    selected = {call.call_id: call for call in calls if call.call_id in split_by_id}
    stages = {
        stage: _stage_results(config, run_id, stage)
        for stage in [
            "classification",
            "agent_performance",
            "patient_experience",
            "safety_compliance",
        ]
    }

    fieldnames = [
        "call_id",
        "split",
        "source_type",
        "stage",
        "metric",
        "draft_status",
        "draft_score",
        "draft_value",
        "draft_reasoning",
        "draft_evidence",
        "human_score",
        "human_value",
        "human_notes",
        "review_status",
        "reviewer",
    ]
    rows: list[dict[str, Any]] = []
    for call_id in calibration_call_ids(manifest):
        call = selected[call_id]
        classification = stages["classification"].get(call_id)
        class_result = classification.result if classification else None
        rows.append(
            {
                "call_id": call_id,
                "split": split_by_id[call_id],
                "source_type": call.source_type.value,
                "stage": "classification",
                "metric": "classification",
                "draft_status": "complete" if class_result else "missing",
                "draft_score": "",
                "draft_value": json.dumps(class_result or {}, ensure_ascii=False),
                "draft_reasoning": (class_result or {}).get("reasoning_summary", ""),
                "draft_evidence": "",
                "human_score": "",
                "human_value": "",
                "human_notes": "",
                "review_status": "pending",
                "reviewer": "",
            }
        )
        for stage, fields in [
            ("agent_performance", AP_FIELDS),
            ("patient_experience", PX_FIELDS),
            ("safety_compliance", SAFETY_FIELDS),
        ]:
            stage_record = stages[stage].get(call_id)
            result = stage_record.result if stage_record and stage_record.result else {}
            for metric in fields:
                metric_result = result.get(metric, {})
                rows.append(
                    {
                        "call_id": call_id,
                        "split": split_by_id[call_id],
                        "source_type": call.source_type.value,
                        "stage": stage,
                        "metric": metric,
                        "draft_status": metric_result.get("status", "missing"),
                        "draft_score": metric_result.get("score", ""),
                        "draft_value": "",
                        "draft_reasoning": metric_result.get("reasoning_summary", ""),
                        "draft_evidence": json.dumps(
                            metric_result.get("evidence", []), ensure_ascii=False
                        ),
                        "human_score": "",
                        "human_value": "",
                        "human_notes": "",
                        "review_status": "pending",
                        "reviewer": "",
                    }
                )
        ap_record = stages["agent_performance"].get(call_id)
        ap_result = ap_record.result if ap_record and ap_record.result else {}
        for metric in ["automation_level_achieved", "best_possible_automation_level"]:
            rows.append(
                {
                    "call_id": call_id,
                    "split": split_by_id[call_id],
                    "source_type": call.source_type.value,
                    "stage": "agent_performance",
                    "metric": metric,
                    "draft_status": "scored" if metric in ap_result else "missing",
                    "draft_score": ap_result.get(metric, ""),
                    "draft_value": "",
                    "draft_reasoning": "Automation level; review against observable outcome.",
                    "draft_evidence": "",
                    "human_score": "",
                    "human_value": "",
                    "human_notes": "",
                    "review_status": "pending",
                    "reviewer": "",
                }
            )

    review_dir = config.path("review") / run_id
    review_dir.mkdir(parents=True, exist_ok=True)
    csv_path = review_dir / "calibration_review.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    labels = [
        HumanLabel(call_id=call_id, split=split_by_id[call_id])
        for call_id in calibration_call_ids(manifest)
    ]
    labels_path = review_dir / "human_labels.jsonl"
    write_jsonl(labels_path, labels)
    return csv_path, labels_path


def _rows_from_csv(path: Path) -> Iterable[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as handle:
        yield from csv.DictReader(handle)


def _rows_from_xlsx(path: Path) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Calibration" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a Calibration sheet")
    sheet = workbook["Calibration"]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(iterator)]
    for values in iterator:
        yield dict(zip(headers, values, strict=False))
    workbook.close()


def import_calibration(config: EvalConfig, run_id: str, source: Path) -> Path:
    if source.suffix.casefold() == ".xlsx":
        rows = _rows_from_xlsx(source)
    elif source.suffix.casefold() == ".csv":
        rows = _rows_from_csv(source)
    else:
        raise ValueError("Human review import must be .csv or .xlsx")

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        call_id = str(row.get("call_id") or "").strip()
        if call_id:
            grouped[call_id].append(dict(row))

    labels: list[HumanLabel] = []
    for call_id, call_rows in grouped.items():
        split = str(call_rows[0].get("split") or "development")
        review_statuses = {str(row.get("review_status") or "pending") for row in call_rows}
        if review_statuses == {"complete"}:
            review_status = "complete"
        elif review_statuses == {"pending"}:
            review_status = "pending"
        else:
            review_status = "in_progress"
        classification: dict[str, Any] | None = None
        ap_scores: dict[str, int | None] = {}
        px_scores: dict[str, int | None] = {}
        safety: dict[str, str] = {}
        automation_level_achieved: int | None = None
        best_possible_automation_level: int | None = None
        notes: list[str] = []
        reviewer = None
        for row in call_rows:
            stage = str(row.get("stage") or "")
            metric = str(row.get("metric") or "")
            reviewer = reviewer or str(row.get("reviewer") or "").strip() or None
            if row.get("human_notes"):
                notes.append(f"{stage}.{metric}: {row['human_notes']}")
            if stage == "classification" and row.get("human_value"):
                classification = json.loads(str(row["human_value"]))
            elif stage in {"agent_performance", "patient_experience"}:
                value = row.get("human_score")
                parsed = int(value) if value not in (None, "") else None
                if stage == "agent_performance" and metric == "automation_level_achieved":
                    automation_level_achieved = parsed
                    continue
                if stage == "agent_performance" and metric == "best_possible_automation_level":
                    best_possible_automation_level = parsed
                    continue
                target = ap_scores if stage == "agent_performance" else px_scores
                target[metric] = parsed
            elif stage == "safety_compliance" and row.get("human_value"):
                safety[metric] = str(row["human_value"])
        labels.append(
            HumanLabel(
                call_id=call_id,
                split=split,
                review_status=review_status,
                reviewer=reviewer,
                classification=classification,
                agent_performance_scores=ap_scores,
                patient_experience_scores=px_scores,
                safety_gate_statuses=safety,
                automation_level_achieved=automation_level_achieved,
                best_possible_automation_level=best_possible_automation_level,
                notes="\n".join(notes),
            )
        )
    output = config.path("runs") / run_id / "human_labels.jsonl"
    write_jsonl(output, sorted(labels, key=lambda label: label.call_id))
    return output
